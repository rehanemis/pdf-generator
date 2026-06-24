import streamlit as st
import openpyxl
import zipfile
import io
import re
from datetime import datetime, date
from pypdf import PdfReader, PdfWriter

# ── Page config ─────────────────────────────────────────────────────────────────
st.set_page_config(page_title="PDF Generator", page_icon="📄", layout="centered")

st.title("📄 Customer PDF Generator")
st.markdown("Upload your Excel file and PDF template to generate one PDF per customer.")

# ── Helper functions ─────────────────────────────────────────────────────────────

def format_date(value, fmt="%d/%m/%Y"):
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime(fmt)
    try:
        for f in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(str(value), f).strftime(fmt)
            except ValueError:
                continue
    except Exception:
        pass
    return str(value)

def generate_ref(prefix, index):
    return f"{prefix}-{str(index).zfill(4)}"

def safe_filename(name):
    keepchars = (".", "_", "-")
    return "".join(c if (c.isalnum() or c in keepchars) else "_" for c in str(name)).strip("_")

def build_filename(pattern, row, ref, headers):
    """
    Build filename from pattern like: {Student Name}_{Grade}_{REF}
    Replaces {ColumnName} with row values, {REF} with reference number.
    """
    result = pattern
    # Replace {REF} first
    result = result.replace("{REF}", ref)
    # Replace any {Column Name} with row value
    for col in headers:
        placeholder = "{" + col + "}"
        if placeholder in result:
            val = str(row.get(col, "")) if row.get(col) is not None else ""
            result = result.replace(placeholder, val)
    # Clean up
    return safe_filename(result) or "output"

def get_excel_rows(excel_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
    ws = wb.active
    headers = [str(cell.value).strip() if cell.value else f"col_{i}"
               for i, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(cell is not None for cell in row):
            rows.append(dict(zip(headers, row)))
    return headers, rows

def fill_pdf_form(pdf_bytes, field_map):
    reader = PdfReader(io.BytesIO(pdf_bytes))
    writer = PdfWriter()
    writer.append(reader)
    writer.update_page_form_field_values(writer.pages[0], field_map)
    out = io.BytesIO()
    writer.write(out)
    out.seek(0)
    return out.read()

def list_pdf_fields(pdf_bytes):
    reader = PdfReader(io.BytesIO(pdf_bytes))
    fields = reader.get_fields()
    return list(fields.keys()) if fields else []

# ── Sidebar settings ─────────────────────────────────────────────────────────────
with st.sidebar:
    st.header("⚙️ Settings")
    ref_prefix = st.text_input("Reference Number Prefix", value="REF-2024")
    date_format = st.selectbox("Date Format in PDF", [
        "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%B %d, %Y", "%d %B %Y"
    ])
    st.markdown("---")
    st.markdown("**Built for:** PDF Automation Workflow")
    st.markdown("**Version:** 2.0")

# ── Step 1: Upload files ─────────────────────────────────────────────────────────
st.header("Step 1: Upload Files")
col1, col2 = st.columns(2)
with col1:
    excel_file = st.file_uploader("📊 Upload Excel File (.xlsx)", type=["xlsx"])
with col2:
    pdf_template = st.file_uploader("📄 Upload PDF Template (.pdf)", type=["pdf"])

# ── Main flow ────────────────────────────────────────────────────────────────────
if excel_file and pdf_template:
    excel_bytes = excel_file.read()
    pdf_bytes = pdf_template.read()

    headers, rows = get_excel_rows(excel_bytes)
    pdf_fields = list_pdf_fields(pdf_bytes)

    st.success(f"✅ Excel loaded: **{len(rows)} rows**, **{len(headers)} columns**")
    if pdf_fields:
        st.success(f"✅ PDF has **{len(pdf_fields)} fillable fields** detected")
    else:
        st.warning("⚠️ No fillable fields found in PDF. Please use a PDF with fillable form fields.")

    # ── Step 2: Preview ──────────────────────────────────────────────────────────
    st.header("Step 2: Preview Data")
    with st.expander("👀 Preview Excel Data (first 5 rows)", expanded=False):
        st.dataframe(rows[:5])

    if pdf_fields:
        # ── Step 3: Field Mapping ────────────────────────────────────────────────
        st.header("Step 3: Map Excel Columns → PDF Fields")
        st.markdown("Match each PDF field to the correct Excel column.")

        col_options = ["-- skip --"] + headers
        field_mapping = {}

        auto_map = {
            "full_name": ["Full Name", "Name", "Customer Name", "Student Name"],
            "first_name": ["First Name", "First"],
            "last_name": ["Last Name", "Surname"],
            "dob": ["Date of Birth", "DOB", "Birth Date"],
            "address": ["Address", "Street"],
            "city": ["City", "Town"],
            "state": ["State", "Province"],
            "zip": ["ZIP", "ZIP Code", "Postal Code"],
            "email": ["Email", "Email Address"],
            "phone": ["Phone", "Mobile"],
        }

        def suggest(pdf_field):
            pf_lower = pdf_field.lower().replace(" ", "_").replace("-", "_")
            for key, candidates in auto_map.items():
                if key in pf_lower:
                    for c in candidates:
                        if c in headers:
                            return c
            for h in headers:
                if h.lower().strip() == pdf_field.lower().strip():
                    return h
            return "-- skip --"

        with st.expander("🔗 Field Mapping", expanded=True):
            cols = st.columns(2)
            for i, field in enumerate(pdf_fields):
                with cols[i % 2]:
                    suggested = suggest(field)
                    default_idx = col_options.index(suggested) if suggested in col_options else 0
                    chosen = st.selectbox(f"PDF: `{field}`", options=col_options,
                                          index=default_idx, key=f"map_{field}")
                    if chosen != "-- skip --":
                        field_mapping[field] = chosen

        # Date columns
        st.markdown("**Which columns are dates?** (will be formatted automatically)")
        date_cols = st.multiselect(
            "Select date columns",
            options=headers,
            default=[h for h in headers if any(d in h.lower() for d in ["date", "dob", "birth", "expiry"])]
        )

        # ── Step 4: Filename Pattern ─────────────────────────────────────────────
        st.header("Step 4: Set Output Filename")

        # Show available placeholders as clickable chips
        st.markdown("**Available placeholders** — click to copy into the pattern:")

        # Show column chips
        chip_cols = st.columns(min(len(headers) + 1, 5))
        all_placeholders = ["{REF}"] + ["{" + h + "}" for h in headers]

        st.markdown(
            " ".join([f"`{p}`" for p in all_placeholders]),
            help="Use these in your filename pattern below"
        )

        # Default pattern suggestion
        first_col = headers[0] if headers else "Name"
        default_pattern = "{" + first_col + "}_{REF}"

        filename_pattern = st.text_input(
            "📝 Filename Pattern",
            value=default_pattern,
            help="Use {ColumnName} for any Excel column, {REF} for reference number"
        )

        # Live preview
        if rows:
            preview_name = build_filename(filename_pattern, rows[0], generate_ref(ref_prefix, 1), headers)
            st.info(f"📄 Preview: **{preview_name}.pdf**")

        # ── Step 5: Generate ─────────────────────────────────────────────────────
        st.header("Step 5: Generate PDFs")
        st.markdown(f"Ready to generate **{len(rows)} PDFs** with **{len(field_mapping)} mapped fields**.")

        if st.button("🚀 Generate All PDFs", type="primary", use_container_width=True):
            if not field_mapping:
                st.error("Please map at least one field before generating.")
            else:
                progress = st.progress(0)
                status = st.empty()
                zip_buffer = io.BytesIO()
                used_names = {}

                with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
                    for i, row in enumerate(rows):
                        ref = generate_ref(ref_prefix, i + 1)

                        # Build field values
                        values = {}
                        for pdf_field, excel_col in field_mapping.items():
                            val = row.get(excel_col, "")
                            if excel_col in date_cols:
                                val = format_date(val, date_format)
                            else:
                                val = str(val) if val is not None else ""
                            values[pdf_field] = val

                        # Auto-fill reference field if exists in PDF
                        for rf in ["reference", "ref", "ref_number", "reference_number"]:
                            if rf in pdf_fields:
                                values[rf] = ref

                        # Fill PDF
                        try:
                            filled = fill_pdf_form(pdf_bytes, values)
                        except Exception as e:
                            st.warning(f"Row {i+1}: Error — {e}")
                            continue

                        # Build filename from pattern
                        base_name = build_filename(filename_pattern, row, ref, headers)

                        # Handle duplicate filenames
                        final_name = base_name
                        if final_name in used_names:
                            used_names[final_name] += 1
                            final_name = f"{base_name}_{used_names[base_name]}"
                        else:
                            used_names[final_name] = 1

                        filename = final_name + ".pdf"
                        zf.writestr(filename, filled)

                        progress.progress((i + 1) / len(rows))
                        status.text(f"Processing {i+1}/{len(rows)}: {filename}")

                zip_buffer.seek(0)
                progress.progress(1.0)
                status.success(f"✅ Done! {len(rows)} PDFs generated successfully.")

                st.download_button(
                    label="📥 Download All PDFs (ZIP)",
                    data=zip_buffer,
                    file_name=f"PDFs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
                    mime="application/zip",
                    use_container_width=True
                )
    else:
        st.info("💡 Your PDF needs fillable form fields. Create them in Adobe Acrobat or LibreOffice.")

else:
    st.info("👆 Please upload both an Excel file and a PDF template to get started.")

st.markdown("---")
st.caption("PDF Generator v2.0 — Upload Excel → Map Fields → Set Filename → Download PDFs")
